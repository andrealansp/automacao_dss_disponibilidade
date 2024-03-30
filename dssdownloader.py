import calendar
import locale
import os
import sys
import time
from datetime import datetime, date
from typing import List
from win32com.client import Dispatch

import openpyxl
from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from emailsender import Emailer
from config import *


def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)


diretorio_downloads: str = resource_path("downloads")
diretorio_disponibilidade: str = resource_path("disponibilidade")
navegador = None


# Remove o arquivo status.xlsx baixado anteriormente
def remove_arquivos():
    try:
        global diretorio_downloads
        # configurando diretorios
        arquivos: List[str] = os.listdir(diretorio_downloads)

        # Remove arquivo anterior
        for arquivo in arquivos:
            os.remove(os.path.join(diretorio_downloads, arquivo))
    except FileNotFoundError:
        print("Sem arquivos na pasta")


# Inicia o driver do navegador
def iniciar_driver():
    chrome_options = Options()

    arguments = ["--lang=en-US", "--window-size=1300,1000", "--headless"]

    for argument in arguments:
        chrome_options.add_argument(argument)

    chrome_options.add_experimental_option(
        "prefs",
        {
            "download.default_directory": diretorio_downloads,
            "safebrowsing.enabled": "false",
            "download.prompt_for_download": False,
            "profile.default_content_setting_values.notifications": 2,
            "profile.default_content_setting_values.automatic_downloads": 1,
        },
    )

    driver = webdriver.Chrome(
        service=ChromeService(ChromeDriverManager().install()), options=chrome_options
    )
    wait = WebDriverWait(
        driver,
        10,
        poll_frequency=1,
        ignored_exceptions=[
            NoSuchElementException,
            ElementNotVisibleException,
            ElementNotSelectableException,
        ],
    )
    return driver, wait


# Realiza o login no site do dss
def logar():
    global navegador
    navegador, wait = iniciar_driver()
    navegador.get("http://189.50.4.158:8281")
    time.sleep(5)
    campo_usuario = navegador.find_element(
        "xpath", "/html/body/div/div/div[3]/div[2]/div[1]/div/input"
    )
    campo_usuario.click()
    campo_usuario.send_keys(LOGIN_DAHUA)
    campo_senha = navegador.find_element(
        "xpath", "/html/body/div/div/div[3]/div[2]/div[2]/div[1]/input"
    )
    campo_senha.click()
    time.sleep(1)
    senha = PASSWORD_DAHUA

    for letra in senha:
        campo_senha.send_keys(letra)
        time.sleep((1 / 5))

    time.sleep(2)
    entrar = navegador.find_element("xpath", "/html/body/div/div/div[3]/div[2]/button")
    entrar.click()
    time.sleep(5)


# Abre a área de dispositivos do site DSS
def acessar_dispositivos():
    link_dispositivo = navegador.find_element(
        "xpath", "/html/body/div/div/div[3]/div/div/div/div/div/div[1]/div/h3"
    )
    link_dispositivo.click()
    time.sleep(5)


# Realiza o download do arquivo.
def realiza_download():
    link_exportado = navegador.find_element(
        "xpath",
        "/html/body/div/div/div[3]/div/div/div[2]/div[2]/div[1]/div[4]/span[2]/div/div[1]/a",
    )
    ActionChains(navegador).move_to_element(link_exportado).perform()
    link_conten_sub = navegador.find_element(
        "xpath",
        '//*[@id="page-content"]/div[2]/div[2]/div[1]/div[4]/span[2]/div/div[2]/div/div/ul/li[1]',
    )

    acoes = ActionChains(navegador)
    acoes.move_to_element(link_conten_sub)
    acoes.click(link_conten_sub)
    acoes.perform()
    time.sleep(20)
    navegador.quit()


# Altera o nome da WorkSheet para status, a planilha original vem um nome composto pela data.
def renomeia_sheet():
    arquivos_baixados: list[str] = os.listdir(diretorio_downloads)
    try:
        for arquivo_baixado in arquivos_baixados:
            os.rename(
                os.path.join(diretorio_downloads, arquivo_baixado),
                os.path.join(diretorio_downloads, "status.xlsx"),
            )

        ss = openpyxl.load_workbook(os.path.join(diretorio_downloads, "status.xlsx"))
        tabela = ss.active
        tabela.title = "status"
        ss.save(os.path.join(diretorio_downloads, "status.xlsx"))
    except Exception as e:
        print(e.args)


# Calcula quantidade de câmeras online e quantidade de câmeras
def calcula_status():
    caminho_status = os.path.join(diretorio_downloads, "status.xlsx")
    status = openpyxl.load_workbook(caminho_status)
    sheet_status = status["status"]
    cameras_online = 0
    quantidade_cameras = 0
    for indice, linha in enumerate(sheet_status.iter_rows(min_row=2), 1):
        cameras_online += int(linha[37].value)
        quantidade_cameras = indice
    return quantidade_cameras, cameras_online


# Atualiza o arquivo de disponibilidade
def atualizar_disponibilidade():
    caminho_disponiblidade = os.path.join(
        diretorio_disponibilidade, "DISPONIBILIDADE_CONTRATO.xlsx"
    )

    disponibilidade = openpyxl.load_workbook(caminho_disponiblidade,data_only=True)

    # definindo variáveis importante
    locale.setlocale(locale.LC_ALL, "pt_br")
    data_de_hoje = datetime.now()
    str_data_de_hoje = data_de_hoje.strftime("%Y-%m-%d 00:00:00")
    mes_numero = data_de_hoje.month
    sheet_mes_atual = disponibilidade.active

    for linha in sheet_mes_atual.iter_rows(min_row=2):               
        data = str(linha[0].value)
        if str_data_de_hoje == data:
            print(linha)
            # Preenchendo as três colunas com os dados calculado na planilha status
            # Colunas  CAMS ONLINE | CAMS TOTAL | PERCENTAGEM DE CÂMERA ONLINE  
            if linha[1].value is None:
                quantidade_cameras, cameras_online = calcula_status()
                linha[1].value = cameras_online
                linha[2].value = quantidade_cameras
                linha[3].value = round(float(cameras_online / quantidade_cameras),4) * 100
                linha[19].value = linha[3].value 
                break
            
            # Colunas  CAMS ONLINE | CAMS TOTAL | PERCENTAGEM DE CÂMERA ONLINE  
            if linha[4].value is None:
                quantidade_cameras, cameras_online = calcula_status()
                linha[4].value = cameras_online
                linha[5].value = quantidade_cameras
                linha[6].value = round(float(cameras_online / quantidade_cameras),4) * 100
                linha[19].value = round(((float(linha[3].value) + float(linha[6].value))/ 2),4)
                break
            
            # Colunas  CAMS ONLINE | CAMS TOTAL | PERCENTAGEM DE CÂMERA ONLINE  
            if linha[7].value is None:
                quantidade_cameras, cameras_online = calcula_status()
                linha[7].value = cameras_online
                linha[8].value = quantidade_cameras
                linha[9].value = round(float(cameras_online / quantidade_cameras),4) * 100
                linha[19].value = round(((float(linha[3].value) + float(linha[6].value) + float(linha[9].value)) / 3 ),4)
                break
            
            # Colunas  CAMS ONLINE | CAMS TOTAL | PERCENTAGEM DE CÂMERA ONLINE  
            if linha[10].value is None:
                quantidade_cameras, cameras_online = calcula_status()
                linha[10].value = cameras_online
                linha[11].value = quantidade_cameras
                linha[12].value = round(float(cameras_online / quantidade_cameras),4) * 100
                linha[19].value = round(((float(linha[3].value) + float(linha[6].value) + 
                                         float(linha[9].value) + float(linha[12].value)) / 4 ),4)
                break
            
            # Colunas  CAMS ONLINE | CAMS TOTAL | PERCENTAGEM DE CÂMERA ONLINE  
            if linha[13].value is None:
                quantidade_cameras, cameras_online = calcula_status()
                linha[13].value = cameras_online
                linha[14].value = quantidade_cameras
                linha[15].value =round(float(cameras_online / quantidade_cameras),4) * 100
                linha[19].value = round(((float(linha[3].value) + float(linha[6].value) + float(linha[9].value) + 
                                         float(linha[12].value) + float(linha[15].value)) / 5 ),4)
                break
            
            # Colunas  CAMS ONLINE | CAMS TOTAL | PERCENTAGEM DE CÂMERA ONLINE 
            if linha[16].value is None:
                quantidade_cameras, cameras_online = calcula_status()
                linha[16].value = cameras_online
                linha[17].value = quantidade_cameras
                linha[18].value =round(float(cameras_online / quantidade_cameras),4) * 100
                linha[19].value = round(((float(linha[3].value) + float(linha[6].value) + float(linha[9].value) + 
                                         float(linha[12].value) + float(linha[15].value) + 
                                         float(linha[18].value))  / 6 ),4)
                break
    
    disponibilidade.save(caminho_disponiblidade)

# Função utilizada para evitar exceção caso tenha dados nulos.
def verifica_se_vazio(media: str) -> str:
    if media is None:
        return "0 %"
    else:
        return f"{round(float(media),2)} %"

# Função para enviar o e-mail
def enviar_email():
    """Realiza o envio de um relatório após as 18:00"""
    
    # Utilizada para configurar o envio de email após as 18:00
    hora_de_corte = datetime.strptime(f"{date.today()} 18:00:00", "%Y-%m-%d %H:%M:%S") 
    hora_atual = datetime.now()

    if hora_atual > hora_de_corte:
        lista_contatos = ["a.alves@perkons.com"]
        email = Emailer(EMAIL_ADDRESS, EMAIL_PASSWORD)
        caminho_disponiblidade = os.path.join(
            diretorio_disponibilidade, "DISPONIBILIDADE 2024.xlsx"
        )
        disponibilidade = openpyxl.load_workbook(caminho_disponiblidade, data_only=True)
        locale.setlocale(locale.LC_ALL, "pt_br")
        data_de_hoje = datetime.now()
        str_data_de_hoje = data_de_hoje.strftime("%Y-%m-%d 00:00:00")
        mes_numero = data_de_hoje.month
        mes_atual = calendar.month_name[mes_numero].upper()
        sheet_mes_atual = disponibilidade[mes_atual]
        mensagem = ""

        for linha in sheet_mes_atual.iter_rows(min_row=2, values_only=True):
            data = str(linha[0])
            if data == str_data_de_hoje:
                mensagem = f"""
                 <style>
                    table {{
                    width: 100%;
                    font-size: 10px;
                    border: 1px gray solid;
                    }}
                    th,
                    td {{
                    border: 1px black solid;
                    height: 50px;
                    text-align: center;
                    }}
                    .content {{
                        padding: 5px;
                        border-radius: 10px;
                        background-color: lightgray;
                        text-align:center;
                    }}
                 </style>
                <div class="content">
                <h1 text-align="center"> Relatório de Disponibilidade DETRAN-ES <h1>
                <table>
                    <thead>
                        <th> Amostra 1</th>
                        <th>Amostra 2</th>
                        <th>Amostra 3</th>
                        <th> Amostra 4</th>
                        <th>Amostra 5</th>
                        <th>Amostra 6</th>
                        <th>Média do Dia</th>
                    </thead>
                    <tbody>
                        <tr>
                            <td>{linha[1]} / {linha[2]} / {verifica_se_vazio(linha[3])}</td>
                            <td>{linha[4]} / {linha[5]} / {verifica_se_vazio(linha[6])}</td>
                            <td>{linha[7]} / {linha[8]} / {verifica_se_vazio(linha[9])}</td>
                            <td>{linha[10]} / {linha[11]} / {verifica_se_vazio(linha[12])}</td>
                            <td>{linha[13]} / {linha[14]} / {verifica_se_vazio(linha[15])}</td>
                            <td>{linha[16]} / {linha[17]} / {verifica_se_vazio(linha[18])}</td>
                            <td>{verifica_se_vazio(linha[19])}</td>                    
                        </tr>
                    </tbody>
                    </table>
                    </div>
                """
        email.definir_conteudo(topico=f"Relatório de Disponibilidade {date.today().strftime("%d/%m/%y")} ",
                           email_remetente="andre@andrealves.eng.br",
                           lista_contatos=lista_contatos,
                           conteudo_email=mensagem)
        try:
            email.enviar_email(intervalo_em_segundos=5)
        except Exception as e:
            print(e.args)
    else:
        # envio de mensagem para motirar se foi realizada corretamente.
        email = Emailer(EMAIL_ADDRESS, EMAIL_PASSWORD)
        lista_contatos = ["a.alves@perkons.com"]
        email.definir_conteudo(topico=f"Relatório de Disponibilidade executado com sucesso ",
                           email_remetente="andre@andrealves.eng.br",
                           lista_contatos=lista_contatos,
                           conteudo_email="Executado com sucesso !")
        try:
            email.enviar_email(intervalo_em_segundos=5)
        except Exception as e:
            print(e.args)


remove_arquivos()
logar()
acessar_dispositivos()
realiza_download()
renomeia_sheet()
atualizar_disponibilidade()
time.sleep(5)
enviar_email()
